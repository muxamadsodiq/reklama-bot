"""Survey / Statistika — super admin UI.

FSM navigatsiyasi:
- SurveyAdmin.q_text  : Yangi savol matnini yozyapti
- SurveyAdmin.q_edit  : Savol matnini tahrirlash
- SurveyAdmin.opt_text: Variant matnini yozyapti (data: question_id, pending_oid? None)
- SurveyAdmin.opt_edit: Variant matnini tahrirlash
- SurveyAdmin.opt_newq_text: Varinatdan chiqadigan YANGI savol matni (data: oid)

Callbacklar (hammasi "sa:st:..." prefiksi bilan, super_admin routeriga tegmaslik uchun):
  sa:st:home              -> Statistika asosiy menyusi
  sa:st:addq              -> Yangi savol matni so'raladi
  sa:st:list              -> Savollar ro'yxati
  sa:st:tree              -> Daraxt ko'rinishi
  sa:st:excel             -> Excel yuklash
  sa:st:clear             -> Javoblarni tozalash (tasdiq so'raydi)
  sa:st:clear:yes         -> Javoblarni haqiqatda tozalash
  sa:st:q:<qid>           -> Bitta savol paneli
  sa:st:q:edit:<qid>      -> Savol matnini tahrirlash
  sa:st:q:del:<qid>       -> Savolni o'chirish
  sa:st:q:root:<qid>      -> Rootga qilish
  sa:st:q:addopt:<qid>    -> Variant qo'shish (matn so'raladi)
  sa:st:opt:<oid>         -> Bitta variant paneli
  sa:st:opt:edit:<oid>    -> Variant matnini tahrirlash
  sa:st:opt:del:<oid>     -> Variantni o'chirish
  sa:st:opt:end:<oid>     -> Tugatish (next=NULL)
  sa:st:opt:pickq:<oid>   -> Mavjud savollar ro'yxatini chiqarish
  sa:st:opt:link:<oid>:<qid> -> Linkni o'rnatish
  sa:st:opt:newq:<oid>    -> Yangi savol yaratib ulash
"""
import io
import logging
from datetime import datetime

from aiogram import Router, F, Bot
from aiogram.filters import Command
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from aiogram.types import (
    Message,
    CallbackQuery,
    InlineKeyboardMarkup,
    InlineKeyboardButton,
    BufferedInputFile,
)

import database as db
from config import SUPER_ADMIN_IDS

router = Router()
log = logging.getLogger(__name__)


async def _sa_only_msg(event: Message) -> bool:
    try:
        if event.chat.type != "private":
            return False
    except Exception:
        return False
    return event.from_user.id in SUPER_ADMIN_IDS


async def _sa_only_cb(event: CallbackQuery) -> bool:
    return event.from_user.id in SUPER_ADMIN_IDS


router.message.filter(_sa_only_msg)
router.callback_query.filter(_sa_only_cb)


class SurveyAdmin(StatesGroup):
    q_text = State()
    q_edit = State()
    opt_text = State()
    opt_edit = State()
    opt_newq_text = State()


def _stat_home_kb():
    return InlineKeyboardMarkup(
        inline_keyboard=[
            [InlineKeyboardButton(text="➕ Yangi savol qo'shish", callback_data="sa:st:addq")],
            [InlineKeyboardButton(text="📋 Savollar ro'yxati", callback_data="sa:st:list")],
            [InlineKeyboardButton(text="🌳 Daraxtni ko'rish", callback_data="sa:st:tree")],
            [InlineKeyboardButton(text="📥 Excelga yuklash", callback_data="sa:st:excel")],
            [InlineKeyboardButton(text="🗑 Javoblarni tozalash", callback_data="sa:st:clear")],
            [InlineKeyboardButton(text="⬅️ Bosh menyu", callback_data="sa:home")],
        ]
    )


async def _render_home(target: Message | CallbackQuery):
    qs = await db.survey_list_questions()
    text = (
        "📊 <b>Statistika / So'rovnoma</b>\n"
        "━━━━━━━━━━━━━━━━━━━\n"
        f"Jami savollar: <b>{len(qs)}</b>\n\n"
        "Daraxtsimon savol-javob tuzamiz. User /start bosganda javob beradi, "
        "keyin Excelga yuklab olasiz."
    )
    kb = _stat_home_kb()
    if isinstance(target, CallbackQuery):
        try:
            await target.message.edit_text(text, reply_markup=kb, parse_mode="HTML")
        except Exception:
            await target.message.answer(text, reply_markup=kb, parse_mode="HTML")
        await target.answer()
    else:
        await target.answer(text, reply_markup=kb, parse_mode="HTML")


@router.callback_query(F.data == "sa:st:home")
async def cb_home(cb: CallbackQuery, state: FSMContext):
    await state.clear()
    await _render_home(cb)


@router.message(Command("statistika"))
async def cmd_stat(msg: Message, state: FSMContext):
    await state.clear()
    await _render_home(msg)


# ============================================================
# YANGI SAVOL QO'SHISH
# ============================================================

@router.callback_query(F.data == "sa:st:addq")
async def cb_addq(cb: CallbackQuery, state: FSMContext):
    await state.set_state(SurveyAdmin.q_text)
    await cb.message.edit_text(
        "✏️ Yangi savol matnini yuboring:\n\n"
        "Masalan: <i>Qayerdansiz?</i>",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[[
            InlineKeyboardButton(text="❌ Bekor qilish", callback_data="sa:st:home")
        ]]),
        parse_mode="HTML",
    )
    await cb.answer()


@router.message(SurveyAdmin.q_text)
async def handle_qtext(msg: Message, state: FSMContext):
    text = (msg.text or "").strip()
    if not text:
        await msg.answer("❗️ Bo'sh bo'lmasin. Qayta yuboring.")
        return
    qid = await db.survey_add_question(text)
    await state.clear()
    await msg.answer(
        f"✅ Savol qo'shildi (#{qid}).\n\nEndi variantlar qo'shing:",
        reply_markup=await _question_panel_kb(qid),
        parse_mode="HTML",
    )
    await _render_question_panel(msg, qid)


# ============================================================
# SAVOLLAR RO'YXATI
# ============================================================

@router.callback_query(F.data == "sa:st:list")
async def cb_list(cb: CallbackQuery, state: FSMContext):
    await state.clear()
    qs = await db.survey_list_questions()
    if not qs:
        await cb.message.edit_text(
            "📋 Savollar yo'q.\nAvval '➕ Yangi savol qo'shish' ni bosing.",
            reply_markup=_stat_home_kb(),
        )
        await cb.answer()
        return
    rows = []
    for q in qs:
        mark = "🌱" if q["is_root"] else "▫️"
        rows.append([InlineKeyboardButton(
            text=f"{mark} #{q['id']} {q['text'][:40]}",
            callback_data=f"sa:st:q:{q['id']}",
        )])
    rows.append([InlineKeyboardButton(text="⬅️ Orqaga", callback_data="sa:st:home")])
    await cb.message.edit_text(
        "📋 <b>Savollar ro'yxati</b>\n\n🌱 — root (birinchi)",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=rows),
        parse_mode="HTML",
    )
    await cb.answer()


# ============================================================
# BITTA SAVOL PANELI
# ============================================================

async def _question_panel_kb(qid: int):
    q = await db.survey_get_question(qid)
    opts = await db.survey_list_options(qid)
    rows = []
    for o in opts:
        if o["next_question_id"]:
            tail = f"→ #{o['next_question_id']}"
        else:
            tail = "🔚"
        rows.append([InlineKeyboardButton(
            text=f"• {o['text'][:30]} {tail}",
            callback_data=f"sa:st:opt:{o['id']}",
        )])
    rows.append([InlineKeyboardButton(text="➕ Variant qo'shish", callback_data=f"sa:st:q:addopt:{qid}")])
    btns2 = [InlineKeyboardButton(text="✏️ Matnni tahrir", callback_data=f"sa:st:q:edit:{qid}")]
    if not q["is_root"]:
        btns2.append(InlineKeyboardButton(text="🌱 Root qilish", callback_data=f"sa:st:q:root:{qid}"))
    rows.append(btns2)
    rows.append([
        InlineKeyboardButton(text="🗑 O'chirish", callback_data=f"sa:st:q:del:{qid}"),
        InlineKeyboardButton(text="⬅️ Ro'yxat", callback_data="sa:st:list"),
    ])
    return InlineKeyboardMarkup(inline_keyboard=rows)


async def _render_question_panel(target, qid: int):
    q = await db.survey_get_question(qid)
    if not q:
        if isinstance(target, CallbackQuery):
            await target.answer("Topilmadi", show_alert=True)
        return
    opts = await db.survey_list_options(qid)
    mark = "🌱 ROOT" if q["is_root"] else "▫️"
    text = f"{mark} <b>#{q['id']}</b>\n\n{q['text']}\n\n"
    if opts:
        text += f"Variantlar ({len(opts)}):\n"
        for o in opts:
            tail = f"→ savol #{o['next_question_id']}" if o["next_question_id"] else "🔚 Tugatish"
            text += f"  • {o['text']}  <i>{tail}</i>\n"
    else:
        text += "<i>Variantlar yo'q. Qo'shing.</i>"
    kb = await _question_panel_kb(qid)
    if isinstance(target, CallbackQuery):
        try:
            await target.message.edit_text(text, reply_markup=kb, parse_mode="HTML")
        except Exception:
            await target.message.answer(text, reply_markup=kb, parse_mode="HTML")
        await target.answer()
    else:
        await target.answer(text, reply_markup=kb, parse_mode="HTML")


@router.callback_query(F.data.regexp(r"^sa:st:q:\d+$"))
async def cb_qpanel(cb: CallbackQuery, state: FSMContext):
    await state.clear()
    qid = int(cb.data.split(":")[3])
    await _render_question_panel(cb, qid)


@router.callback_query(F.data.startswith("sa:st:q:edit:"))
async def cb_qedit(cb: CallbackQuery, state: FSMContext):
    qid = int(cb.data.split(":")[4])
    await state.set_state(SurveyAdmin.q_edit)
    await state.update_data(qid=qid)
    await cb.message.edit_text(
        "✏️ Yangi matnni yuboring:",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[[
            InlineKeyboardButton(text="❌ Bekor", callback_data=f"sa:st:q:{qid}")
        ]]),
    )
    await cb.answer()


@router.message(SurveyAdmin.q_edit)
async def handle_qedit(msg: Message, state: FSMContext):
    data = await state.get_data()
    qid = data.get("qid")
    text = (msg.text or "").strip()
    if not text:
        await msg.answer("❗️ Bo'sh bo'lmasin.")
        return
    await db.survey_update_question(qid, text)
    await state.clear()
    await msg.answer("✅ Yangilandi.")
    await _render_question_panel(msg, qid)


@router.callback_query(F.data.startswith("sa:st:q:del:"))
async def cb_qdel(cb: CallbackQuery, state: FSMContext):
    qid = int(cb.data.split(":")[4])
    await db.survey_delete_question(qid)
    await cb.answer("O'chirildi", show_alert=False)
    await cb_list(cb, state)


@router.callback_query(F.data.startswith("sa:st:q:root:"))
async def cb_qroot(cb: CallbackQuery):
    qid = int(cb.data.split(":")[4])
    await db.survey_set_root(qid)
    await cb.answer("🌱 Root qilindi")
    await _render_question_panel(cb, qid)


# ============================================================
# VARIANT QO'SHISH
# ============================================================

@router.callback_query(F.data.startswith("sa:st:q:addopt:"))
async def cb_addopt(cb: CallbackQuery, state: FSMContext):
    qid = int(cb.data.split(":")[4])
    await state.set_state(SurveyAdmin.opt_text)
    await state.update_data(qid=qid)
    await cb.message.edit_text(
        f"✏️ Savol #{qid} uchun variant matnini yuboring:\n\nMasalan: <i>Toshkent</i>",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[[
            InlineKeyboardButton(text="❌ Bekor", callback_data=f"sa:st:q:{qid}")
        ]]),
        parse_mode="HTML",
    )
    await cb.answer()


@router.message(SurveyAdmin.opt_text)
async def handle_opttext(msg: Message, state: FSMContext):
    data = await state.get_data()
    qid = data.get("qid")
    text = (msg.text or "").strip()
    if not text:
        await msg.answer("❗️ Bo'sh bo'lmasin.")
        return
    oid = await db.survey_add_option(qid, text, next_qid=None)
    await state.clear()
    await msg.answer(
        f"✅ Variant qo'shildi.\n\nBu variant bosilganda nima bo'lsin?",
        reply_markup=_option_link_kb(oid),
    )


def _option_link_kb(oid: int):
    return InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="🔚 Tugatish (yakuniy)", callback_data=f"sa:st:opt:end:{oid}")],
        [InlineKeyboardButton(text="➡️ Mavjud savolga ulash", callback_data=f"sa:st:opt:pickq:{oid}")],
        [InlineKeyboardButton(text="🆕 Yangi savol yaratib ulash", callback_data=f"sa:st:opt:newq:{oid}")],
    ])


# ============================================================
# BITTA VARIANT PANELI
# ============================================================

async def _render_option_panel(target: CallbackQuery, oid: int):
    o = await db.survey_get_option(oid)
    if not o:
        await target.answer("Topilmadi", show_alert=True)
        return
    q = await db.survey_get_question(o["question_id"])
    next_txt = f"savol #{o['next_question_id']}" if o["next_question_id"] else "🔚 Tugatish"
    text = (
        f"▫️ Variant #{o['id']}\n"
        f"Savol: <b>{q['text']}</b>\n\n"
        f"Matn: <b>{o['text']}</b>\n"
        f"Keyingi: <b>{next_txt}</b>"
    )
    kb = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="✏️ Matnni tahrir", callback_data=f"sa:st:opt:edit:{oid}")],
        [InlineKeyboardButton(text="🔗 Ulashni o'zgartirish", callback_data=f"sa:st:opt:pickq:{oid}")],
        [InlineKeyboardButton(text="🔚 Tugatish (yakuniy)", callback_data=f"sa:st:opt:end:{oid}")],
        [InlineKeyboardButton(text="🆕 Yangi savol yaratib ulash", callback_data=f"sa:st:opt:newq:{oid}")],
        [InlineKeyboardButton(text="🗑 O'chirish", callback_data=f"sa:st:opt:del:{oid}")],
        [InlineKeyboardButton(text="⬅️ Savol paneliga", callback_data=f"sa:st:q:{o['question_id']}")],
    ])
    try:
        await target.message.edit_text(text, reply_markup=kb, parse_mode="HTML")
    except Exception:
        await target.message.answer(text, reply_markup=kb, parse_mode="HTML")
    await target.answer()


@router.callback_query(F.data.regexp(r"^sa:st:opt:\d+$"))
async def cb_optpanel(cb: CallbackQuery, state: FSMContext):
    await state.clear()
    oid = int(cb.data.split(":")[3])
    await _render_option_panel(cb, oid)


@router.callback_query(F.data.startswith("sa:st:opt:edit:"))
async def cb_optedit(cb: CallbackQuery, state: FSMContext):
    oid = int(cb.data.split(":")[4])
    await state.set_state(SurveyAdmin.opt_edit)
    await state.update_data(oid=oid)
    await cb.message.edit_text("✏️ Yangi matnni yuboring:",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[[
            InlineKeyboardButton(text="❌ Bekor", callback_data=f"sa:st:opt:{oid}")
        ]]))
    await cb.answer()


@router.message(SurveyAdmin.opt_edit)
async def handle_optedit(msg: Message, state: FSMContext):
    data = await state.get_data()
    oid = data.get("oid")
    text = (msg.text or "").strip()
    if not text:
        await msg.answer("❗️ Bo'sh bo'lmasin.")
        return
    await db.survey_update_option(oid, text)
    await state.clear()
    o = await db.survey_get_option(oid)
    await msg.answer("✅ Yangilandi.")
    await _render_question_panel(msg, o["question_id"])


@router.callback_query(F.data.startswith("sa:st:opt:del:"))
async def cb_optdel(cb: CallbackQuery):
    oid = int(cb.data.split(":")[4])
    o = await db.survey_get_option(oid)
    qid = o["question_id"] if o else None
    await db.survey_delete_option(oid)
    await cb.answer("O'chirildi")
    if qid:
        await _render_question_panel(cb, qid)
    else:
        await _render_home(cb)


@router.callback_query(F.data.startswith("sa:st:opt:end:"))
async def cb_optend(cb: CallbackQuery):
    oid = int(cb.data.split(":")[4])
    await db.survey_set_option_next(oid, None)
    await cb.answer("🔚 Tugatish o'rnatildi")
    o = await db.survey_get_option(oid)
    await _render_question_panel(cb, o["question_id"])


@router.callback_query(F.data.startswith("sa:st:opt:pickq:"))
async def cb_optpickq(cb: CallbackQuery):
    oid = int(cb.data.split(":")[4])
    qs = await db.survey_list_questions()
    o = await db.survey_get_option(oid)
    cur_q = o["question_id"]
    rows = []
    for q in qs:
        if q["id"] == cur_q:
            continue  # o'ziga ulanmasin (loop)
        rows.append([InlineKeyboardButton(
            text=f"#{q['id']} {q['text'][:35]}",
            callback_data=f"sa:st:opt:link:{oid}:{q['id']}",
        )])
    rows.append([InlineKeyboardButton(text="⬅️ Orqaga", callback_data=f"sa:st:opt:{oid}")])
    if not qs or len(rows) == 1:
        await cb.answer("Boshqa savollar yo'q", show_alert=True)
        return
    await cb.message.edit_text("Qaysi savolga ulaymiz?", reply_markup=InlineKeyboardMarkup(inline_keyboard=rows))
    await cb.answer()


@router.callback_query(F.data.startswith("sa:st:opt:link:"))
async def cb_optlink(cb: CallbackQuery):
    parts = cb.data.split(":")
    oid, qid = int(parts[4]), int(parts[5])
    await db.survey_set_option_next(oid, qid)
    await cb.answer(f"🔗 Savol #{qid} ga ulandi")
    await _render_option_panel(cb, oid)


@router.callback_query(F.data.startswith("sa:st:opt:newq:"))
async def cb_optnewq(cb: CallbackQuery, state: FSMContext):
    oid = int(cb.data.split(":")[4])
    await state.set_state(SurveyAdmin.opt_newq_text)
    await state.update_data(oid=oid)
    await cb.message.edit_text(
        "✏️ Yangi savol matnini yuboring.\n\nU yaratilib, shu variantga ulanadi.",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[[
            InlineKeyboardButton(text="❌ Bekor", callback_data=f"sa:st:opt:{oid}")
        ]]),
    )
    await cb.answer()


@router.message(SurveyAdmin.opt_newq_text)
async def handle_optnewq(msg: Message, state: FSMContext):
    data = await state.get_data()
    oid = data.get("oid")
    text = (msg.text or "").strip()
    if not text:
        await msg.answer("❗️ Bo'sh bo'lmasin.")
        return
    new_qid = await db.survey_add_question(text)
    await db.survey_set_option_next(oid, new_qid)
    await state.clear()
    await msg.answer(
        f"✅ Savol #{new_qid} yaratildi va ulandi.\n\nEndi bu yangi savolga variantlar qo'shing:"
    )
    await _render_question_panel(msg, new_qid)


# ============================================================
# DARAXT KO'RINISHI
# ============================================================

@router.callback_query(F.data == "sa:st:tree")
async def cb_tree(cb: CallbackQuery, state: FSMContext):
    await state.clear()
    qs = await db.survey_list_questions()
    if not qs:
        await cb.answer("Savollar yo'q", show_alert=True)
        return
    # Rootdan boshlab BFS bilan chizamiz, qolganlari 'ORPHAN' bo'limida
    root = await db.survey_get_root_question()
    lines = []
    visited = set()

    async def render_q(qid: int, depth: int):
        if qid in visited:
            lines.append("  " * depth + f"↺ (sikl: #{qid})")
            return
        visited.add(qid)
        q = await db.survey_get_question(qid)
        if not q:
            return
        prefix = "🌱 " if q["is_root"] else ""
        lines.append("  " * depth + f"{prefix}#{q['id']}: {q['text']}")
        opts = await db.survey_list_options(qid)
        for o in opts:
            if o["next_question_id"]:
                lines.append("  " * (depth + 1) + f"• {o['text']} →")
                await render_q(o["next_question_id"], depth + 2)
            else:
                lines.append("  " * (depth + 1) + f"• {o['text']} 🔚")

    if root:
        await render_q(root["id"], 0)

    orphans = [q for q in qs if q["id"] not in visited]
    if orphans:
        lines.append("")
        lines.append("⚠️ <b>Ulanmagan savollar:</b>")
        for q in orphans:
            lines.append(f"  #{q['id']}: {q['text']}")

    text = "🌳 <b>Daraxt</b>\n\n<pre>" + "\n".join(lines[:200]) + "</pre>"
    await cb.message.edit_text(
        text,
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[[
            InlineKeyboardButton(text="⬅️ Orqaga", callback_data="sa:st:home")
        ]]),
        parse_mode="HTML",
    )
    await cb.answer()


# ============================================================
# JAVOBLARNI TOZALASH
# ============================================================

@router.callback_query(F.data == "sa:st:clear")
async def cb_clear(cb: CallbackQuery):
    await cb.message.edit_text(
        "⚠️ Hamma javoblarni va sessionlarni o'chirasizmi?\n(Savollar qoladi)",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="✅ Ha, tozalash", callback_data="sa:st:clear:yes")],
            [InlineKeyboardButton(text="⬅️ Bekor", callback_data="sa:st:home")],
        ]),
    )
    await cb.answer()


@router.callback_query(F.data == "sa:st:clear:yes")
async def cb_clear_yes(cb: CallbackQuery):
    await db.survey_clear_all_answers()
    await cb.answer("🗑 Tozalandi", show_alert=True)
    await _render_home(cb)


# ============================================================
# EXCEL YUKLASH
# ============================================================

@router.callback_query(F.data == "sa:st:excel")
async def cb_excel(cb: CallbackQuery, bot: Bot):
    await cb.answer("📥 Tayyorlanyapti...")
    try:
        from openpyxl import Workbook
    except ImportError:
        await cb.message.answer("❌ openpyxl o'rnatilmagan. `pip install openpyxl` kerak.")
        return

    data = await db.survey_export_rows()
    questions = data["questions"]
    rows = data["rows"]
    detail = data.get("detail", [])

    wb = Workbook()
    ws = wb.active
    ws.title = "Javoblar"
    # Header
    header = ["user_id", "username", "full_name", "phone", "tugatgan"]
    for q in questions:
        header.append(f"#{q['id']}: {q['text']}")
    ws.append(header)
    # Rows
    for r in rows:
        row = [
            r["user_id"],
            r["username"] or "",
            r["full_name"] or "",
            r["phone"] or "",
            r["completed"],
        ]
        for q in questions:
            row.append(r["answers"].get(q["id"], ""))
        ws.append(row)
    # Width
    from openpyxl.utils import get_column_letter
    widths = [14, 18, 22, 18, 10] + [28] * len(questions)
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

    # 2-sheet: Batafsil (har qator = 1 javob)
    ws2 = wb.create_sheet("Batafsil")
    ws2.append(["user_id", "username", "full_name", "phone", "savol", "javob", "vaqt"])
    for d in detail:
        ws2.append([
            d["user_id"], d["username"], d["full_name"], d["phone"],
            d["question_text"], d["option_text"], d["answered_at"],
        ])
    for i, w in enumerate([14, 18, 22, 18, 40, 28, 22], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    ws2.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"survey_results_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
    await cb.message.answer_document(
        BufferedInputFile(buf.read(), filename=fname),
        caption=f"📥 {len(rows)} ta user · {len(questions)} ta savol · {len(detail)} ta javob",
    )
